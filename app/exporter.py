import io
import csv
import html
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from typing import List, Dict, Any

class ResultExporter:
    @staticmethod
    def export_to_excel(comparison_data: Dict[str, Any], list_configs: List[Dict[str, Any]]) -> bytes:
        """
        Genera un archivo Excel (.xlsx) completo con estilos, colores para el precio más barato,
        y hojas de 'Comparación Completa', 'Resumen y Totales' y 'Top Oportunidades'.
        """
        wb = openpyxl.Workbook()
        
        # Estilos generales
        font_title = Font(name="Arial", size=14, bold=True, color="1E293B")
        font_subtitle = Font(name="Arial", size=11, bold=True, color="334155")
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=9, bold=True)
        font_data = Font(name="Arial", size=9)
        
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        fill_cheapest = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Verde claro
        fill_warning = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amarillo claro
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        border_thin = Side(style="thin", color="CBD5E1")
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        
        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # -------------------------------------------------------------
        # HOJA 1: Comparación Completa (100% de las filas)
        # -------------------------------------------------------------
        ws_comp = wb.active
        ws_comp.title = "Comparación Completa"
        
        # Título
        ws_comp.merge_cells("A1:K1")
        ws_comp["A1"] = "COMPARACIÓN DE LISTAS DE PRECIOS - DETALLE COMPLETO"
        ws_comp["A1"].font = font_title
        ws_comp["A1"].alignment = align_left
        ws_comp.row_dimensions[1].height = 28
        
        p1_name = list_configs[0]['nombre'] if len(list_configs) > 0 else "Lista 1"
        p2_name = list_configs[1]['nombre'] if len(list_configs) > 1 else "Lista 2"
        p3_name = list_configs[2]['nombre'] if len(list_configs) > 2 else "Lista 3"
        
        headers = [
            "Producto / Descripción",
            "Código",
            "Marca",
            "Presentación",
            "Cant.",
            f"Precio {p1_name}",
            f"Precio {p2_name}",
            f"Precio {p3_name}",
            "Proveedor Más Barato",
            "Diferencia ($)",
            "Diferencia (%)",
            "Estado"
        ]
        
        ws_comp.row_dimensions[3].height = 24
        for col_num, h in enumerate(headers, 1):
            cell = ws_comp.cell(row=3, column=col_num, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = cell_border

        rows = comparison_data.get('rows', [])
        for row_idx, r in enumerate(rows, 4):
            ws_comp.row_dimensions[row_idx].height = 20
            
            p1_val = r.get('precio_l1')
            p2_val = r.get('precio_l2')
            p3_val = r.get('precio_l3')
            min_p = r.get('precio_min')
            
            p1_display = f"${p1_val:,.2f}" if p1_val is not None else "No disponible"
            p2_display = f"${p2_val:,.2f}" if p2_val is not None else "No disponible"
            p3_display = f"${p3_val:,.2f}" if p3_val is not None else "No disponible"
            
            diff_money = f"${r['diferencia_dinero']:,.2f}" if r.get('diferencia_dinero') is not None else "$0.00"
            diff_pct = f"{r['diferencia_porcentaje']:.2f}%" if r.get('diferencia_porcentaje') is not None else "0.00%"
            
            row_values = [
                r.get('producto', ''),
                r.get('codigo', ''),
                r.get('marca', ''),
                r.get('presentacion', ''),
                r.get('cantidad', 1),
                p1_display,
                p2_display,
                p3_display,
                r.get('proveedor_mas_barato', ''),
                diff_money,
                diff_pct,
                r.get('estado_precio', '')
            ]
            
            is_even = (row_idx % 2 == 0)
            
            for col_num, val in enumerate(row_values, 1):
                cell = ws_comp.cell(row=row_idx, column=col_num, value=val)
                cell.font = font_data
                cell.border = cell_border
                
                # Alineación
                if col_num in (1, 3, 4, 9, 12):
                    cell.alignment = align_left
                elif col_num in (2, 5):
                    cell.alignment = align_center
                else:
                    cell.alignment = align_right
                    
                # Fondo alternado
                if is_even:
                    cell.fill = fill_zebra
                    
                # Resaltar precios más baratos en verde
                if col_num == 6 and p1_val is not None and min_p is not None and abs(p1_val - min_p) < 0.001 and (p2_val is not None or p3_val is not None):
                    cell.fill = fill_cheapest
                    cell.font = font_bold
                elif col_num == 7 and p2_val is not None and min_p is not None and abs(p2_val - min_p) < 0.001 and (p1_val is not None or p3_val is not None):
                    cell.fill = fill_cheapest
                    cell.font = font_bold
                elif col_num == 8 and p3_val is not None and min_p is not None and abs(p3_val - min_p) < 0.001 and (p1_val is not None or p2_val is not None):
                    cell.fill = fill_cheapest
                    cell.font = font_bold

        # Autoajustar anchos de columna
        for col in ws_comp.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_comp.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        ws_comp.column_dimensions['A'].width = 35 # Producto más ancho

        # -------------------------------------------------------------
        # HOJA 2: Resumen y Totales Financieros
        # -------------------------------------------------------------
        ws_tot = wb.create_sheet(title="Resumen y Totales")
        totals = comparison_data.get('totals', {})
        
        ws_tot["A1"] = "INFORME EJECUTIVO DE COMPARACIÓN DE PRECIOS"
        ws_tot["A1"].font = font_title
        ws_tot.row_dimensions[1].height = 28
        
        ws_tot["A3"] = "1. Resumen de Totales y Ahorros"
        ws_tot["A3"].font = font_subtitle
        
        headers_tot = ["Proveedor / Escenario", "Total Productos Comparables", "Total General Lista", "Ahorro Posible ($)", "Ahorro Posible (%)", "Prod. Más Baratos", "Prod. Exclusivos"]
        for col_num, h in enumerate(headers_tot, 1):
            cell = ws_tot.cell(row=4, column=col_num, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = cell_border
            
        row_cur = 5
        tot_comp_dict = totals.get('totales_comparables', {})
        tot_gen_dict = totals.get('totales_generales', {})
        ahorros_dict = totals.get('ahorros', {})
        baratos_dict = totals.get('conteo_mas_baratos', {})
        excl_dict = totals.get('conteo_exclusivos', {})
        
        for i, cfg in enumerate(list_configs):
            p_name = cfg['nombre']
            tot_c = tot_comp_dict.get(i, 0.0)
            tot_g = tot_gen_dict.get(i, 0.0)
            ahorro_info = ahorros_dict.get(i, {})
            ah_m = ahorro_info.get('ahorro_dinero', 0.0)
            ah_p = ahorro_info.get('ahorro_porcentaje', 0.0)
            cnt_b = baratos_dict.get(i, 0)
            cnt_e = excl_dict.get(i, 0)
            
            data_row = [
                p_name,
                f"${tot_c:,.2f}",
                f"${tot_g:,.2f}",
                f"${ah_m:,.2f}",
                f"{ah_p:.2f}%",
                cnt_b,
                cnt_e
            ]
            for col_num, val in enumerate(data_row, 1):
                cell = ws_tot.cell(row=row_cur, column=col_num, value=val)
                cell.font = font_data
                cell.border = cell_border
                cell.alignment = align_left if col_num == 1 else align_right
            row_cur += 1
            
        # Fila Canasta Óptima (Mejor combinación de precios)
        opt_comp = totals.get('total_optimo_comparables', 0.0)
        opt_total = totals.get('total_compra_optima', 0.0)
        opt_row = ["CANASTA ÓPTIMA (Comprando al más barato)", f"${opt_comp:,.2f}", f"${opt_total:,.2f}", "-", "-", "-", "-"]
        for col_num, val in enumerate(opt_row, 1):
            cell = ws_tot.cell(row=row_cur, column=col_num, value=val)
            cell.font = font_bold
            cell.fill = fill_cheapest
            cell.border = cell_border
            cell.alignment = align_left if col_num == 1 else align_right
            
        # Métricas Globales
        row_cur += 3
        ws_tot.cell(row=row_cur, column=1, value="2. Estadísticas Generales").font = font_subtitle
        row_cur += 1
        
        metrics = [
            ("Total de productos procesados (100% de filas):", totals.get('total_productos_comparados', 0)),
            ("Productos comparables (presentes en 2 o 3 listas):", totals.get('total_productos_comparables', 0)),
            ("Productos exclusivos de un solo proveedor:", totals.get('total_productos_exclusivos', 0)),
            ("Productos con precio idéntico en todos los proveedores:", totals.get('total_precios_iguales', 0)),
            ("Productos sin precio válido (precio $0 o nulo):", totals.get('total_sin_precio_valido', 0)),
            ("Diferencia total monetaria entre listas:", f"${totals.get('diferencia_total_listas', 0.0):,.2f}")
        ]
        
        for label, val in metrics:
            c1 = ws_tot.cell(row=row_cur, column=1, value=label)
            c2 = ws_tot.cell(row=row_cur, column=2, value=val)
            c1.font = font_data
            c2.font = font_bold
            row_cur += 1
            
        for col in ws_tot.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_tot.column_dimensions[col_letter].width = max(max_len + 4, 18)

        # -------------------------------------------------------------
        # HOJA 3: Top Oportunidades de Ahorro
        # -------------------------------------------------------------
        ws_top = wb.create_sheet(title="Top Dispersión de Precios")
        ws_top["A1"] = "PRODUCTOS CON MAYOR DIFERENCIA DE PRECIO (OPORTUNIDADES DE AHORRO)"
        ws_top["A1"].font = font_title
        
        top_headers = ["Producto", "Código", "Precio Más Barato", "Precio Más Caro", "Proveedor Más Barato", "Diferencia ($)", "Diferencia (%)"]
        for col_num, h in enumerate(top_headers, 1):
            cell = ws_top.cell(row=3, column=col_num, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = cell_border
            
        top_items = totals.get('top_diferencia_porcentaje', [])
        for row_idx, item in enumerate(top_items, 4):
            row_data = [
                item.get('producto', ''),
                item.get('codigo', ''),
                f"${item.get('precio_min', 0.0):,.2f}",
                f"${item.get('precio_max', 0.0):,.2f}",
                item.get('proveedor_mas_barato', ''),
                f"${item.get('diferencia_dinero', 0.0):,.2f}",
                f"{item.get('diferencia_porcentaje', 0.0):.2f}%"
            ]
            for col_num, val in enumerate(row_data, 1):
                cell = ws_top.cell(row=row_idx, column=col_num, value=val)
                cell.font = font_data
                cell.border = cell_border
                cell.alignment = align_left if col_num in (1, 5) else (align_center if col_num == 2 else align_right)

        for col in ws_top.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_top.column_dimensions[col_letter].width = max(max_len + 4, 15)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def export_to_csv(comparison_data: Dict[str, Any], list_configs: List[Dict[str, Any]]) -> bytes:
        """Exporta el 100% de los productos comparados a CSV delimitado por comas con BOM UTF-8."""
        out = io.StringIO()
        writer = csv.writer(out, delimiter=';')
        
        p1_name = list_configs[0]['nombre'] if len(list_configs) > 0 else "Lista 1"
        p2_name = list_configs[1]['nombre'] if len(list_configs) > 1 else "Lista 2"
        p3_name = list_configs[2]['nombre'] if len(list_configs) > 2 else "Lista 3"
        
        headers = [
            "Producto", "Codigo", "Marca", "Presentacion", "Cantidad",
            f"Precio_{p1_name}", f"Precio_{p2_name}", f"Precio_{p3_name}",
            "Proveedor_Mas_Barato", "Diferencia_Dinero", "Diferencia_Porcentaje", "Estado"
        ]
        writer.writerow(headers)
        
        for r in comparison_data.get('rows', []):
            writer.writerow([
                r.get('producto', ''),
                r.get('codigo', ''),
                r.get('marca', ''),
                r.get('presentacion', ''),
                r.get('cantidad', 1),
                r.get('precio_l1', '') if r.get('precio_l1') is not None else '',
                r.get('precio_l2', '') if r.get('precio_l2') is not None else '',
                r.get('precio_l3', '') if r.get('precio_l3') is not None else '',
                r.get('proveedor_mas_barato', ''),
                r.get('diferencia_dinero', 0.0),
                r.get('diferencia_porcentaje', 0.0),
                r.get('estado_precio', '')
            ])
            
        # Retornar con BOM UTF-8 para compatibilidad en Excel
        return ('\ufeff' + out.getvalue()).encode('utf-8')

    @staticmethod
    def export_to_pdf(comparison_data: Dict[str, Any], list_configs: List[Dict[str, Any]]) -> bytes:
        """Genera un reporte ejecutivo en PDF apaisado con métricas y tabla comparativa."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            leftMargin=0.4 * inch,
            rightMargin=0.4 * inch,
            topMargin=0.4 * inch,
            bottomMargin=0.4 * inch
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#1E293B"),
            spaceAfter=10
        )
        sub_style = ParagraphStyle(
            name="SubStyle",
            parent=styles["Heading2"],
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#334155"),
            spaceAfter=8
        )
        cell_style = ParagraphStyle(
            name="CellStyle",
            parent=styles["Normal"],
            fontSize=7,
            leading=9
        )
        cell_bold = ParagraphStyle(
            name="CellBold",
            parent=styles["Normal"],
            fontSize=7,
            leading=9,
            fontName="Helvetica-Bold"
        )
        
        story = []
        
        # Título
        story.append(Paragraph("Informe Comparativo de Precios de Proveedores", title_style))
        story.append(Paragraph("Resumen Ejecutivo y Métricas de Ahorro", sub_style))
        story.append(Spacer(1, 10))
        
        # Tabla de Totales
        totals = comparison_data.get('totals', {})
        tot_comp_dict = totals.get('totales_comparables', {})
        tot_gen_dict = totals.get('totales_generales', {})
        ahorros_dict = totals.get('ahorros', {})
        baratos_dict = totals.get('conteo_mas_baratos', {})
        
        summary_table_data = [
            [
                Paragraph("<b>Proveedor</b>", cell_bold),
                Paragraph("<b>Total Comparables</b>", cell_bold),
                Paragraph("<b>Total General</b>", cell_bold),
                Paragraph("<b>Ahorro Posible ($)</b>", cell_bold),
                Paragraph("<b>Ahorro (%)</b>", cell_bold),
                Paragraph("<b>Ganador (Prod. Más Baratos)</b>", cell_bold)
            ]
        ]
        
        for i, cfg in enumerate(list_configs):
            p_name = cfg['nombre']
            tot_c = tot_comp_dict.get(i, 0.0)
            tot_g = tot_gen_dict.get(i, 0.0)
            ah_m = ahorros_dict.get(i, {}).get('ahorro_dinero', 0.0)
            ah_p = ahorros_dict.get(i, {}).get('ahorro_porcentaje', 0.0)
            cnt_b = baratos_dict.get(i, 0)
            
            summary_table_data.append([
                Paragraph(p_name, cell_style),
                Paragraph(f"${tot_c:,.2f}", cell_style),
                Paragraph(f"${tot_g:,.2f}", cell_style),
                Paragraph(f"${ah_m:,.2f}", cell_style),
                Paragraph(f"{ah_p:.2f}%", cell_style),
                Paragraph(f"{cnt_b} artículos", cell_style)
            ])
            
        # Canasta Óptima
        opt_comp = totals.get('total_optimo_comparables', 0.0)
        opt_total = totals.get('total_compra_optima', 0.0)
        summary_table_data.append([
            Paragraph("<b>CANASTA ÓPTIMA (Mínimos)</b>", cell_bold),
            Paragraph(f"<b>${opt_comp:,.2f}</b>", cell_bold),
            Paragraph(f"<b>${opt_total:,.2f}</b>", cell_bold),
            Paragraph("<b>MÁXIMO AHORRO</b>", cell_bold),
            Paragraph("-", cell_style),
            Paragraph("100% de artículos", cell_style)
        ])
        
        t_summary = Table(summary_table_data, colWidths=[1.8*inch, 1.4*inch, 1.4*inch, 1.4*inch, 1.0*inch, 1.8*inch])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E293B")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#DCFCE7")), # Verde canasta optima
        ]))
        story.append(t_summary)
        story.append(Spacer(1, 15))
        
        # Tabla de Detalles (Primeros 100 productos más representativos o destacados en el PDF)
        story.append(Paragraph(f"Detalle de Productos Comparados (Mostrando muestra de {min(len(comparison_data.get('rows', [])), 150)} de {len(comparison_data.get('rows', []))} productos)", sub_style))
        
        p1_name = list_configs[0]['nombre'] if len(list_configs) > 0 else "L1"
        p2_name = list_configs[1]['nombre'] if len(list_configs) > 1 else "L2"
        p3_name = list_configs[2]['nombre'] if len(list_configs) > 2 else "L3"
        
        detail_table_data = [
            [
                Paragraph("<b>Producto</b>", cell_bold),
                Paragraph("<b>Código</b>", cell_bold),
                Paragraph(f"<b>{p1_name}</b>", cell_bold),
                Paragraph(f"<b>{p2_name}</b>", cell_bold),
                Paragraph(f"<b>{p3_name}</b>", cell_bold),
                Paragraph("<b>Más Barato</b>", cell_bold),
                Paragraph("<b>Dif. $</b>", cell_bold),
                Paragraph("<b>Dif. %</b>", cell_bold)
            ]
        ]
        
        for r in comparison_data.get('rows', [])[:150]:
            p1_str = f"${r['precio_l1']:,.2f}" if r.get('precio_l1') is not None else "-"
            p2_str = f"${r['precio_l2']:,.2f}" if r.get('precio_l2') is not None else "-"
            p3_str = f"${r['precio_l3']:,.2f}" if r.get('precio_l3') is not None else "-"
            
            prod_name = html.escape(str(r.get('producto', '') or '')[:35])
            cod_name = html.escape(str(r.get('codigo', '') or '')[:12])
            prov_best = html.escape(str(r.get('proveedor_mas_barato', '') or '')[:15])

            detail_table_data.append([
                Paragraph(prod_name, cell_style),
                Paragraph(cod_name, cell_style),
                Paragraph(p1_str, cell_style),
                Paragraph(p2_str, cell_style),
                Paragraph(p3_str, cell_style),
                Paragraph(prov_best, cell_style),
                Paragraph(f"${r.get('diferencia_dinero', 0.0):,.2f}", cell_style),
                Paragraph(f"{r.get('diferencia_porcentaje', 0.0):.2f}%", cell_style)
            ])
            
        t_detail = Table(detail_table_data, colWidths=[2.2*inch, 0.9*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.3*inch, 0.8*inch, 0.7*inch])
        t_detail.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#334155")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")])
        ]))
        story.append(t_detail)
        
        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def export_purchase_order_excel(comparison_data: Dict[str, Any], list_configs: List[Dict[str, Any]], target_idx: int) -> bytes:
        """
        Genera una planilla de Orden de Compra en Excel para enviar a un proveedor específico,
        incluyendo únicamente los artículos donde ese proveedor es el más económico (o exclusivo).
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        prov_name = list_configs[target_idx]['nombre'] if target_idx < len(list_configs) else f"Proveedor {target_idx+1}"
        ws.title = f"Pedido {prov_name[:20]}"
        
        font_title = Font(name="Arial", size=14, bold=True, color="1E293B")
        font_subtitle = Font(name="Arial", size=10, color="64748B")
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_data = Font(name="Arial", size=9)
        font_bold = Font(name="Arial", size=10, bold=True)
        
        fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        fill_total = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        border_thin = Side(style="thin", color="CBD5E1")
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        
        # Título
        ws.merge_cells("A1:G1")
        ws["A1"] = f"ORDEN DE COMPRA SUGERIDA - {prov_name.upper()}"
        ws["A1"].font = font_title
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26
        
        ws.merge_cells("A2:G2")
        ws["A2"] = "Lista de artículos seleccionados por ser la opción más económica o exclusiva."
        ws["A2"].font = font_subtitle
        ws.row_dimensions[2].height = 18
        
        headers = ["#", "Código", "Producto / Descripción", "Marca", "Presentación", "Cantidad", "Precio Unitario", "Subtotal"]
        ws.row_dimensions[4].height = 22
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border
            
        rows = comparison_data.get('rows', [])
        # Filtrar los artículos ganadores de este proveedor
        order_items = [r for r in rows if r.get('proveedor_mas_barato_idx') == target_idx]
        
        price_key = f"precio_l{target_idx+1}"
        code_key = f"cod_l{target_idx+1}"
        desc_key = f"desc_l{target_idx+1}"
        
        total_order = 0.0
        for i, item in enumerate(order_items, 5):
            ws.row_dimensions[i].height = 19
            p_unit = item.get(price_key) or item.get('precio_min') or 0.0
            qty = item.get('cantidad', 1) or 1
            subtot = round(p_unit * qty, 2)
            total_order += subtot
            
            row_vals = [
                i - 4,
                item.get(code_key) or item.get('codigo', ''),
                item.get(desc_key) or item.get('producto', ''),
                item.get('marca', ''),
                item.get('presentacion', ''),
                qty,
                f"${p_unit:,.2f}",
                f"${subtot:,.2f}"
            ]
            
            is_even = (i % 2 == 0)
            for col_num, val in enumerate(row_vals, 1):
                cell = ws.cell(row=i, column=col_num, value=val)
                cell.font = font_data
                cell.border = cell_border
                if is_even: cell.fill = fill_zebra
                
                if col_num in (1, 6): cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in (7, 8): cell.alignment = Alignment(horizontal="right", vertical="center")
                else: cell.alignment = Alignment(horizontal="left", vertical="center")
                
        # Fila de Total
        tot_row = len(order_items) + 5
        ws.row_dimensions[tot_row].height = 24
        ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=7)
        c_lbl = ws.cell(row=tot_row, column=1, value=f"TOTAL SUGERIDO A COMPRAR A {prov_name.upper()}:")
        c_lbl.font = font_bold
        c_lbl.alignment = Alignment(horizontal="right", vertical="center")
        c_lbl.fill = fill_total
        
        c_val = ws.cell(row=tot_row, column=8, value=f"${total_order:,.2f}")
        c_val.font = font_bold
        c_val.alignment = Alignment(horizontal="right", vertical="center")
        c_val.fill = fill_total
        c_val.border = cell_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        ws.column_dimensions['C'].width = 38

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

